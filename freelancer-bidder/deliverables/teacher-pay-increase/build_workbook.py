"""
Teacher Pay Increase Spreadsheet
Bid project: Teacher Pay Increase Spreadsheet Development

Usage:
  pip install openpyxl
  python build_workbook.py
Output: teacher_pay_increase.xlsx
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUT = Path(__file__).with_name("teacher_pay_increase.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
CALC_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

SAMPLE_TEACHERS = [
    ("T001", "Alice Brown", "Math", 5, 42000),
    ("T002", "Brian Clark", "English", 12, 51000),
    ("T003", "Carla Diaz", "Science", 2, 38000),
    ("T004", "David Evans", "History", 8, 46000),
    ("T005", "Elena Foster", "Art", 15, 54000),
    ("T006", "Frank Green", "PE", 3, 39000),
    ("T007", "Grace Hill", "Math", 9, 47500),
    ("T008", "Henry Ives", "English", 1, 36000),
    ("T009", "Ivy Jones", "Science", 7, 45000),
    ("T010", "Jack King", "Music", 11, 50000),
    ("T011", "Kara Lee", "Math", 4, 41000),
    ("T012", "Leo Martin", "English", 6, 43500),
    ("T013", "Mia Nelson", "Science", 14, 53000),
    ("T014", "Noah Owen", "History", 2, 37500),
    ("T015", "Olivia Perez", "Art", 10, 49000),
    ("T016", "Paul Quinn", "PE", 8, 45500),
    ("T017", "Quinn Reed", "Math", 3, 39500),
    ("T018", "Rita Shah", "English", 13, 52000),
    ("T019", "Sam Turner", "Science", 5, 42500),
    ("T020", "Tina Underwood", "Music", 7, 44500),
    ("T021", "Uma Vargas", "History", 9, 47000),
    ("T022", "Victor White", "Math", 16, 56000),
    ("T023", "Wendy Xu", "English", 4, 40500),
    ("T024", "Xander Young", "Science", 6, 44000),
    ("T025", "Yara Zane", "Art", 2, 37000),
    ("T026", "Zoe Adams", "PE", 11, 49500),
    ("T027", "Amy Blake", "Math", 8, 46500),
    ("T028", "Ben Cole", "English", 1, 35500),
    ("T029", "Cara Dunn", "Science", 10, 48500),
    ("T030", "Dan Ellis", "History", 5, 41500),
]


def style_header(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def autosize(ws, min_width: int = 12, max_width: int = 28) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = max(min_width, min(max_width, length + 2))


def build() -> Path:
    wb = Workbook()

    # --- Config ---
    cfg = wb.active
    cfg.title = "Config"
    cfg["A1"] = "Pay Increase Settings"
    cfg["A1"].font = Font(bold=True, size=14)
    cfg["A3"] = "Mode"
    cfg["B3"] = "tiered"  # flat | tiered
    cfg["B3"].fill = INPUT_FILL
    cfg["C3"] = "flat = use Flat Increase % for everyone; tiered = use Years of Service bands"
    cfg["A4"] = "Flat Increase %"
    cfg["B4"] = 0.05
    cfg["B4"].number_format = "0.00%"
    cfg["B4"].fill = INPUT_FILL
    cfg["A6"] = "Tier"
    cfg["B6"] = "Min Years"
    cfg["C6"] = "Max Years"
    cfg["D6"] = "Increase %"
    style_header(cfg, 6, 4)
    tiers = [
        (1, 0, 2, 0.03),
        (2, 3, 5, 0.05),
        (3, 6, 10, 0.07),
        (4, 11, 99, 0.09),
    ]
    for i, (tier, mn, mx, pct) in enumerate(tiers, start=7):
        cfg.cell(i, 1, tier)
        cfg.cell(i, 2, mn).fill = INPUT_FILL
        cfg.cell(i, 3, mx).fill = INPUT_FILL
        cell = cfg.cell(i, 4, pct)
        cell.number_format = "0.00%"
        cell.fill = INPUT_FILL
        for c in range(1, 5):
            cfg.cell(i, c).border = THIN
    cfg["A12"] = "How to use"
    cfg["A12"].font = Font(bold=True)
    cfg["A13"] = "1) Edit yellow cells in Config"
    cfg["A14"] = "2) Fill / replace teachers on Teachers sheet (yellow = input)"
    cfg["A15"] = "3) Green columns calculate automatically"
    cfg["A16"] = "4) Open Summary for totals"
    autosize(cfg)

    mode_dv = DataValidation(type="list", formula1='"flat,tiered"', allow_blank=False)
    cfg.add_data_validation(mode_dv)
    mode_dv.add(cfg["B3"])

    # --- Teachers ---
    ws = wb.create_sheet("Teachers", 0)
    headers = [
        "Teacher ID",
        "Name",
        "Subject",
        "Years of Service",
        "Current Annual Pay",
        "Increase % Applied",
        "Increase Amount",
        "New Annual Pay",
        "Monthly New Pay",
        "Validation",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(1, col, h)
    style_header(ws, 1, len(headers))
    ws.row_dimensions[1].height = 32

    for i, (tid, name, subject, years, pay) in enumerate(SAMPLE_TEACHERS, start=2):
        ws.cell(i, 1, tid).fill = INPUT_FILL
        ws.cell(i, 2, name).fill = INPUT_FILL
        ws.cell(i, 3, subject).fill = INPUT_FILL
        ws.cell(i, 4, years).fill = INPUT_FILL
        pay_cell = ws.cell(i, 5, pay)
        pay_cell.fill = INPUT_FILL
        pay_cell.number_format = '"$"#,##0.00'

        # Increase %: flat or tiered via Config
        pct_formula = (
            f'=IF(Config!$B$3="flat",Config!$B$4,'
            f'IF(D{i}="","",'
            f'IF(AND(D{i}>=Config!$B$7,D{i}<=Config!$C$7),Config!$D$7,'
            f'IF(AND(D{i}>=Config!$B$8,D{i}<=Config!$C$8),Config!$D$8,'
            f'IF(AND(D{i}>=Config!$B$9,D{i}<=Config!$C$9),Config!$D$9,'
            f'IF(AND(D{i}>=Config!$B$10,D{i}<=Config!$C$10),Config!$D$10,0))))))'
        )
        pct_cell = ws.cell(i, 6, pct_formula)
        pct_cell.number_format = "0.00%"
        pct_cell.fill = CALC_FILL

        amt_cell = ws.cell(i, 7, f"=IF(OR(E{i}=\"\",F{i}=\"\"),\"\",E{i}*F{i})")
        amt_cell.number_format = '"$"#,##0.00'
        amt_cell.fill = CALC_FILL

        new_cell = ws.cell(i, 8, f"=IF(E{i}=\"\",\"\",E{i}+G{i})")
        new_cell.number_format = '"$"#,##0.00'
        new_cell.fill = CALC_FILL

        month_cell = ws.cell(i, 9, f"=IF(H{i}=\"\",\"\",H{i}/12)")
        month_cell.number_format = '"$"#,##0.00'
        month_cell.fill = CALC_FILL

        ws.cell(
            i,
            10,
            f'=IF(A{i}="","",IF(OR(D{i}<0,E{i}<0),"Check years/pay","OK"))',
        ).fill = CALC_FILL

        for c in range(1, 11):
            ws.cell(i, c).border = THIN

    # Pre-format extra empty input rows (31-50) for more teachers
    for i in range(32, 51):
        for c in range(1, 6):
            ws.cell(i, c).fill = INPUT_FILL
            ws.cell(i, c).border = THIN
        ws.cell(i, 6, f'=IF(A{i}="","",IF(Config!$B$3="flat",Config!$B$4,IF(AND(D{i}>=Config!$B$7,D{i}<=Config!$C$7),Config!$D$7,IF(AND(D{i}>=Config!$B$8,D{i}<=Config!$C$8),Config!$D$8,IF(AND(D{i}>=Config!$B$9,D{i}<=Config!$C$9),Config!$D$9,IF(AND(D{i}>=Config!$B$10,D{i}<=Config!$C$10),Config!$D$10,0))))))')
        ws.cell(i, 6).number_format = "0.00%"
        ws.cell(i, 6).fill = CALC_FILL
        ws.cell(i, 7, f'=IF(OR(A{i}="",E{i}=""),"",E{i}*F{i})')
        ws.cell(i, 7).number_format = '"$"#,##0.00'
        ws.cell(i, 7).fill = CALC_FILL
        ws.cell(i, 8, f'=IF(A{i}="","",E{i}+G{i})')
        ws.cell(i, 8).number_format = '"$"#,##0.00'
        ws.cell(i, 8).fill = CALC_FILL
        ws.cell(i, 9, f'=IF(A{i}="","",H{i}/12)')
        ws.cell(i, 9).number_format = '"$"#,##0.00'
        ws.cell(i, 9).fill = CALC_FILL
        ws.cell(i, 10, f'=IF(A{i}="","",IF(OR(D{i}<0,E{i}<0),"Check years/pay","OK"))')
        ws.cell(i, 10).fill = CALC_FILL
        for c in range(6, 11):
            ws.cell(i, c).border = THIN

    autosize(ws, min_width=14, max_width=22)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{len(SAMPLE_TEACHERS) + 1}"

    # --- Summary ---
    sm = wb.create_sheet("Summary")
    sm["A1"] = "Payroll Increase Summary"
    sm["A1"].font = Font(bold=True, size=14)
    sm["A3"] = "Teacher count"
    sm["B3"] = '=COUNTA(Teachers!A2:A50)-COUNTBLANK(Teachers!A2:A50)'
    sm["A4"] = "Current payroll (annual)"
    sm["B4"] = "=SUMIF(Teachers!A2:A50,\"<>\",Teachers!E2:E50)"
    sm["B4"].number_format = '"$"#,##0.00'
    sm["A5"] = "Total increase (annual)"
    sm["B5"] = "=SUMIF(Teachers!A2:A50,\"<>\",Teachers!G2:G50)"
    sm["B5"].number_format = '"$"#,##0.00'
    sm["A6"] = "New payroll (annual)"
    sm["B6"] = "=SUMIF(Teachers!A2:A50,\"<>\",Teachers!H2:H50)"
    sm["B6"].number_format = '"$"#,##0.00'
    sm["A7"] = "New payroll (monthly)"
    sm["B7"] = "=B6/12"
    sm["B7"].number_format = '"$"#,##0.00'
    sm["A9"] = "Validation issues"
    sm["B9"] = '=COUNTIF(Teachers!J2:J50,"Check years/pay")'
    for r in range(3, 10):
        sm.cell(r, 1).font = Font(bold=True)
    autosize(sm)

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    path = build()
    print(f"Created: {path}")
