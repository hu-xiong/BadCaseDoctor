"""Sample preview only — not the full research deliverable."""

from __future__ import annotations

from pathlib import Path

from fastapi import FastAPI
from fastapi.responses import FileResponse, HTMLResponse
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "telangana_irrigation_SAMPLE.xlsx"

app = FastAPI(title="Telangana Irrigation Dataset — SAMPLE")


def load_rows() -> list[dict]:
    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h or "") for h in rows[0]]
    out = []
    for row in rows[1:]:
        if not any(row):
            continue
        out.append({headers[i]: ("" if row[i] is None else str(row[i])) for i in range(len(headers))})
    return out


@app.get("/download")
def download():
    return FileResponse(
        XLSX,
        filename="telangana_irrigation_SAMPLE.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/", response_class=HTMLResponse)
def index():
    rows = load_rows()
    headers = list(rows[0].keys()) if rows else []
    thead = "".join(f"<th>{h}</th>" for h in headers)
    tbody = "".join(
        "<tr>" + "".join(f"<td>{r.get(h, '')}</td>" for h in headers) + "</tr>" for r in rows
    )
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Telangana Irrigation Dataset — SAMPLE Preview</title>
  <style>
    body {{ margin:0; font-family:Segoe UI,system-ui,sans-serif; background:#0f1419; color:#e7eef8; }}
    .wrap {{ max-width:1200px; margin:0 auto; padding:28px 18px 50px; }}
    h1 {{ margin:0 0 8px; font-size:1.45rem; }}
    .banner {{
      background:#3a2612; color:#ffc36a; border:1px solid #6a4a20; border-radius:10px;
      padding:12px 14px; margin:14px 0 18px; line-height:1.45;
    }}
    .sub {{ color:#8b9bb4; margin-bottom:16px; line-height:1.45; }}
    a.btn {{
      display:inline-block; background:#3d9cfd; color:#061018; text-decoration:none;
      font-weight:700; padding:10px 14px; border-radius:8px; margin-bottom:16px;
    }}
    .card {{ background:#1a2332; border:1px solid #2a3648; border-radius:12px; padding:12px; overflow:auto; }}
    table {{ width:100%; border-collapse:collapse; font-size:0.82rem; }}
    th, td {{ border-bottom:1px solid #2a3648; padding:8px; text-align:left; vertical-align:top; }}
    th {{ color:#8b9bb4; white-space:nowrap; }}
    td {{ min-width:110px; }}
  </style>
</head>
<body>
  <div class="wrap">
    <h1>Telangana Irrigation Projects Dataset</h1>
    <div class="banner">
      <strong>SAMPLE PREVIEW ONLY</strong> — {len(rows)} example rows with public citations to validate column layout
      (constituency grouping + chronology + source URL + remarks). This is <em>not</em> the full statewide research file.
      After award I expand to coverage with dual-source checks.
    </div>
    <p class="sub">Columns match your brief: project name, constituency, sanction/completion, ministers, citation, remarks.</p>
    <a class="btn" href="/download">Download SAMPLE .xlsx</a>
    <div class="card">
      <table>
        <thead><tr>{thead}</tr></thead>
        <tbody>{tbody}</tbody>
      </table>
    </div>
  </div>
</body>
</html>
"""
