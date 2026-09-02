"""
Tecnostamp product scraper -> Excel
Bid project: Web Scraping – Extract All Products from website into Excel

Usage:
  pip install -r requirements.txt
  python scrape.py
  python scrape.py --base https://www.tecnostamp.it/en/ --out tecnostamp_products.xlsx
"""

from __future__ import annotations

import argparse
import re
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

DEFAULT_BASE = "https://www.tecnostamp.it/en/"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}


@dataclass
class Product:
    name: str = ""
    sku: str = ""
    category: str = ""
    price: str = ""
    description: str = ""
    url: str = ""
    image_url: str = ""


def session() -> requests.Session:
    s = requests.Session()
    s.headers.update(HEADERS)
    return s


def get_soup(s: requests.Session, url: str) -> BeautifulSoup:
    r = s.get(url, timeout=30)
    r.raise_for_status()
    if "fireshield.js" in r.text and len(r.text) < 2500:
        raise RuntimeError(
            "FireShield bot protection detected. Use scrape_playwright.py instead "
            "(pip install playwright && playwright install chromium)."
        )
    return BeautifulSoup(r.text, "lxml")


def same_site(url: str, base: str) -> bool:
    return urlparse(url).netloc == urlparse(base).netloc


def discover_seed_urls(s: requests.Session, base: str) -> list[str]:
    soup = get_soup(s, base)
    seeds = {base.rstrip("/") + "/"}
    for a in soup.select("a[href]"):
        href = urljoin(base, a.get("href", ""))
        if not same_site(href, base):
            continue
        path = urlparse(href).path.lower()
        # catalog / product listing style paths
        if any(
            k in path
            for k in (
                "product",
                "catalog",
                "shop",
                "category",
                "collection",
                "stamp",
                "punch",
                "die",
            )
        ):
            seeds.add(href.split("#")[0])
    return sorted(seeds)


def extract_product_links(soup: BeautifulSoup, page_url: str, base: str) -> set[str]:
    links: set[str] = set()
    selectors = [
        "a.product-item-link",
        "a.woocommerce-LoopProduct-link",
        ".product a[href]",
        ".products a[href]",
        "a[href*='/product']",
        "a[href*='/products/']",
        "a[href*='prodotto']",
    ]
    for sel in selectors:
        for a in soup.select(sel):
            href = urljoin(page_url, a.get("href", ""))
            if same_site(href, base):
                links.add(href.split("#")[0])
    # pagination
    for a in soup.select("a.next, a[rel='next'], .pagination a[href]"):
        href = urljoin(page_url, a.get("href", ""))
        if same_site(href, base):
            links.add(href.split("#")[0])
    return links


def looks_like_product(url: str, soup: BeautifulSoup) -> bool:
    path = urlparse(url).path.lower()
    if any(x in path for x in ("/product/", "/products/", "/prodotto/")):
        return True
    if soup.select_one("meta[property='og:type'][content*='product']"):
        return True
    if soup.select_one(".product-info, .product_title, h1.product_title, .sku"):
        return True
    return False


def parse_product(url: str, soup: BeautifulSoup) -> Product:
    name = ""
    for sel in ("h1.product_title", "h1.page-title", "h1", "meta[property='og:title']"):
        el = soup.select_one(sel)
        if not el:
            continue
        name = el.get("content") if el.name == "meta" else el.get_text(" ", strip=True)
        if name:
            break

    sku = ""
    for sel in (".sku", "[itemprop='sku']", ".product-sku"):
        el = soup.select_one(sel)
        if el:
            sku = el.get_text(" ", strip=True)
            break
    if not sku:
        m = re.search(r"SKU[:\s]*([A-Za-z0-9\-_./]+)", soup.get_text(" ", strip=True), re.I)
        if m:
            sku = m.group(1)

    price = ""
    for sel in (".price .amount", ".woocommerce-Price-amount", "[itemprop='price']", ".price"):
        el = soup.select_one(sel)
        if el:
            price = el.get("content") or el.get_text(" ", strip=True)
            if price:
                break

    desc = ""
    for sel in (
        ".woocommerce-product-details__short-description",
        "#tab-description",
        ".product-description",
        "meta[property='og:description']",
    ):
        el = soup.select_one(sel)
        if el:
            desc = el.get("content") if el.name == "meta" else el.get_text(" ", strip=True)
            if desc:
                break
    desc = re.sub(r"\s+", " ", desc)[:2000]

    category = " > ".join(
        a.get_text(" ", strip=True)
        for a in soup.select(".posted_in a, .breadcrumb a, .breadcrumbs a")
        if a.get_text(strip=True)
    )

    image = ""
    for sel in (
        "meta[property='og:image']",
        ".woocommerce-product-gallery__image img",
        ".product-image img",
        "img.wp-post-image",
    ):
        el = soup.select_one(sel)
        if not el:
            continue
        image = el.get("content") or el.get("src") or el.get("data-src") or ""
        if image:
            image = urljoin(url, image)
            break

    return Product(
        name=name,
        sku=sku,
        category=category,
        price=price,
        description=desc,
        url=url,
        image_url=image,
    )


def crawl(base: str, delay: float = 0.6, max_pages: int = 400) -> list[Product]:
    s = session()
    seeds = discover_seed_urls(s, base)
    queue = list(seeds)
    seen_pages: set[str] = set()
    products: dict[str, Product] = {}

    print(f"Seeds ({len(seeds)}):")
    for u in seeds[:20]:
        print(f"  - {u}")

    while queue and len(seen_pages) < max_pages:
        url = queue.pop(0)
        if url in seen_pages:
            continue
        seen_pages.add(url)
        try:
            soup = get_soup(s, url)
        except Exception as exc:
            print(f"[warn] {url}: {exc}")
            continue

        if looks_like_product(url, soup):
            prod = parse_product(url, soup)
            if prod.name:
                products[url] = prod
                print(f"[product] {prod.name[:80]}")

        for link in extract_product_links(soup, url, base):
            if link not in seen_pages and link not in queue:
                # keep crawl focused
                path = urlparse(link).path.lower()
                if any(
                    k in path
                    for k in (
                        "product",
                        "catalog",
                        "shop",
                        "category",
                        "collection",
                        "stamp",
                        "page",
                        "prodotto",
                    )
                ) or link.rstrip("/") == base.rstrip("/"):
                    queue.append(link)

        time.sleep(delay)

    return list(products.values())


def write_excel(rows: list[Product], out: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    headers = ["Name", "SKU", "Category", "Price", "Description", "URL", "Image URL"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
    for p in rows:
        d = asdict(p)
        ws.append(
            [
                d["name"],
                d["sku"],
                d["category"],
                d["price"],
                d["description"],
                d["url"],
                d["image_url"],
            ]
        )
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
    wb.save(out)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base", default=DEFAULT_BASE)
    parser.add_argument("--out", default="tecnostamp_products.xlsx")
    parser.add_argument("--delay", type=float, default=0.6)
    parser.add_argument("--max-pages", type=int, default=400)
    args = parser.parse_args()

    try:
        products = crawl(args.base, delay=args.delay, max_pages=args.max_pages)
    except RuntimeError as exc:
        print(f"[error] {exc}")
        raise SystemExit(2) from exc
    out = Path(args.out)
    write_excel(products, out)
    print(f"\nWrote {len(products)} products -> {out.resolve()}")


if __name__ == "__main__":
    main()
