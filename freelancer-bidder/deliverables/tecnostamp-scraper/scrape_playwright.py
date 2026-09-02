"""
Tecnostamp scraper via Playwright (bypasses FireShield JS challenge).

Usage:
  pip install playwright openpyxl
  playwright install chromium
  python scrape_playwright.py --out tecnostamp_products.xlsx
"""

from __future__ import annotations

import argparse
import re
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from urllib.parse import urljoin, urlparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

DEFAULT_BASE = "https://www.tecnostamp.it/en/"


@dataclass
class Product:
    name: str = ""
    sku: str = ""
    category: str = ""
    price: str = ""
    description: str = ""
    url: str = ""
    image_url: str = ""


def write_excel(rows: list[Product], out: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    headers = ["Name", "SKU", "Category", "Price", "Description", "URL", "Image URL"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
    for p in rows:
        d = asdict(p)
        ws.append([d[k] for k in ("name", "sku", "category", "price", "description", "url", "image_url")])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
    wb.save(out)


def crawl(base: str, max_pages: int = 250, headless: bool = True) -> list[Product]:
    from playwright.sync_api import sync_playwright

    products: dict[str, Product] = {}
    queue = [base]
    seen: set[str] = set()
    host = urlparse(base).netloc

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="en-US",
        )
        page = context.new_page()

        while queue and len(seen) < max_pages:
            url = queue.pop(0)
            if url in seen:
                continue
            seen.add(url)
            print(f"[page] {url}")
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=60000)
                page.wait_for_timeout(2500)
            except Exception as exc:
                print(f"[warn] goto failed: {exc}")
                continue

            html = page.content()
            if "fireshield.js" in html and len(html) < 2000:
                print("[info] FireShield challenge — waiting longer…")
                page.wait_for_timeout(8000)
                html = page.content()

            # collect links
            hrefs = page.eval_on_selector_all(
                "a[href]",
                "els => els.map(e => e.href)",
            )
            for href in hrefs:
                if not href or urlparse(href).netloc != host:
                    continue
                href = href.split("#")[0]
                path = urlparse(href).path.lower()
                if href not in seen and any(
                    k in path
                    for k in ("product", "catalog", "shop", "category", "collection", "stamp", "prodotto")
                ):
                    queue.append(href)

            # parse if product-like
            title = page.title()
            h1 = ""
            try:
                h1 = page.locator("h1").first.inner_text(timeout=1000).strip()
            except Exception:
                pass
            path = urlparse(url).path.lower()
            is_product = any(k in path for k in ("/product", "/prodotto")) or bool(
                page.locator(".sku, .product_title, .woocommerce-Price-amount").count()
            )
            if is_product and (h1 or title):
                sku = ""
                price = ""
                desc = ""
                image = ""
                try:
                    sku = page.locator(".sku").first.inner_text(timeout=500).strip()
                except Exception:
                    pass
                try:
                    price = page.locator(".price").first.inner_text(timeout=500).strip()
                except Exception:
                    pass
                try:
                    desc = page.locator(
                        ".woocommerce-product-details__short-description, #tab-description"
                    ).first.inner_text(timeout=500)
                    desc = re.sub(r"\s+", " ", desc)[:2000]
                except Exception:
                    pass
                try:
                    image = page.locator("meta[property='og:image']").first.get_attribute("content") or ""
                except Exception:
                    pass
                products[url] = Product(
                    name=h1 or title,
                    sku=sku,
                    price=price,
                    description=desc,
                    url=url,
                    image_url=image,
                )
                print(f"  + product: {(h1 or title)[:80]}")

            time.sleep(0.4)

        browser.close()
    return list(products.values())


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base", default=DEFAULT_BASE)
    parser.add_argument("--out", default="tecnostamp_products.xlsx")
    parser.add_argument("--max-pages", type=int, default=250)
    parser.add_argument("--headed", action="store_true")
    args = parser.parse_args()

    rows = crawl(args.base, max_pages=args.max_pages, headless=not args.headed)
    out = Path(args.out)
    write_excel(rows, out)
    print(f"\nWrote {len(rows)} products -> {out.resolve()}")


if __name__ == "__main__":
    main()
