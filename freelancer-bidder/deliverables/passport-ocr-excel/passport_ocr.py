"""
Indian Passport OCR -> Excel
Bid project: Passport OCR to Excel Tool

Columns (required order):
  Name, Date of Birth, Passport Number, Issue Date, Expiry Date,
  Place of Birth, Nationality, Gender

Usage:
  pip install -r requirements.txt
  # Install Tesseract OCR for Windows and ensure `tesseract` is on PATH
  # Optional: Poppler for PDF (pdf2image)
  python passport_ocr.py path/to/file_or_folder [--out results.xlsx] [--preview]

GUI:
  python passport_ocr.py --gui
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import asdict, dataclass, fields
from pathlib import Path
from typing import Iterable

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image

try:
    import pytesseract
except ImportError:  # pragma: no cover
    pytesseract = None

try:
    from pdf2image import convert_from_path
except ImportError:  # pragma: no cover
    convert_from_path = None


IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"}
PDF_EXTS = {".pdf"}


@dataclass
class PassportRecord:
    name: str = ""
    date_of_birth: str = ""
    passport_number: str = ""
    issue_date: str = ""
    expiry_date: str = ""
    place_of_birth: str = ""
    nationality: str = ""
    gender: str = ""
    source_file: str = ""
    confidence_note: str = ""


FIELD_HEADERS = [
    "Name",
    "Date of Birth",
    "Passport Number",
    "Issue Date",
    "Expiry Date",
    "Place of Birth",
    "Nationality",
    "Gender",
]


def preprocess(image: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    gray = cv2.bilateralFilter(gray, 9, 75, 75)
    thr = cv2.adaptiveThreshold(
        gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, 11
    )
    return thr


def pil_to_bgr(img: Image.Image) -> np.ndarray:
    rgb = np.array(img.convert("RGB"))
    return cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR)


def load_images(path: Path) -> list[tuple[str, np.ndarray]]:
    path = path.resolve()
    if path.is_dir():
        files = sorted(
            p
            for p in path.iterdir()
            if p.suffix.lower() in IMAGE_EXTS | PDF_EXTS and p.is_file()
        )
    else:
        files = [path]

    out: list[tuple[str, np.ndarray]] = []
    for f in files:
        ext = f.suffix.lower()
        if ext in IMAGE_EXTS:
            data = cv2.imdecode(np.fromfile(str(f), dtype=np.uint8), cv2.IMREAD_COLOR)
            if data is None:
                print(f"[warn] cannot read image: {f}", file=sys.stderr)
                continue
            out.append((str(f), data))
        elif ext in PDF_EXTS:
            if convert_from_path is None:
                print(f"[warn] pdf2image not installed, skip PDF: {f}", file=sys.stderr)
                continue
            try:
                pages = convert_from_path(str(f), dpi=300)
            except Exception as exc:  # pragma: no cover
                print(f"[warn] PDF convert failed ({f}): {exc}", file=sys.stderr)
                continue
            for idx, page in enumerate(pages, start=1):
                out.append((f"{f}#page{idx}", pil_to_bgr(page)))
    return out


def ocr_text(image: np.ndarray) -> str:
    if pytesseract is None:
        raise RuntimeError("pytesseract is not installed")
    proc = preprocess(image)
    config = "--oem 3 --psm 6"
    text = pytesseract.image_to_string(proc, lang="eng", config=config)
    return text.replace("\r", "\n")


def _find(patterns: Iterable[str], text: str, flags: int = re.I) -> str:
    for pat in patterns:
        m = re.search(pat, text, flags)
        if m:
            return (m.group(1) if m.lastindex else m.group(0)).strip(" :-\t")
    return ""


def normalize_date(raw: str) -> str:
    raw = raw.strip().upper().replace(" ", "")
    # DD/MM/YYYY or DD-MM-YYYY
    m = re.search(r"(\d{2})[/-](\d{2})[/-](\d{4})", raw)
    if m:
        return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
    # DD MMM YYYY
    m = re.search(
        r"(\d{2})([A-Z]{3})(\d{4})",
        raw.replace(" ", ""),
    )
    if m:
        return f"{m.group(1)} {m.group(2)} {m.group(3)}"
    return raw


def parse_indian_passport(text: str, source: str) -> PassportRecord:
    clean = re.sub(r"[ \t]+", " ", text)
    upper = clean.upper()

    passport_number = _find(
        [
            r"\b([A-Z][0-9]{7})\b",
            r"PASSPORT\s*NO\.?\s*[:\-]?\s*([A-Z0-9]{8})",
        ],
        upper,
    )

    date_pat = (
        r"([0-9]{2}[/\-][0-9]{2}[/\-][0-9]{4}"
        r"|[0-9]{2}\s*[A-Z]{3}\s*[0-9]{4}"
        r"|[0-9]{2}[A-Z]{3}[0-9]{4})"
    )
    dob = normalize_date(
        _find(
            [
                rf"DATE OF BIRTH[:\s\-]*{date_pat}",
                rf"\bDOB[:\s\-]*{date_pat}",
            ],
            upper,
        )
    )
    issue = normalize_date(
        _find(
            [
                rf"DATE OF ISSUE[:\s\-]*{date_pat}",
                rf"\bDOI[:\s\-]*{date_pat}",
            ],
            upper,
        )
    )
    expiry = normalize_date(
        _find(
            [
                rf"DATE OF EXPIR(?:Y|ATION)[:\s\-]*{date_pat}",
                rf"\bDOE[:\s\-]*{date_pat}",
            ],
            upper,
        )
    )

    gender = _find(
        [
            r"\bSEX[:\s\-]*([MF])\b",
            r"\bGENDER[:\s\-]*(MALE|FEMALE|M|F)\b",
        ],
        upper,
    )
    if gender in {"MALE", "M"}:
        gender = "M"
    elif gender in {"FEMALE", "F"}:
        gender = "F"

    nationality = _find(
        [
            r"NATIONALITY[:\s\-]*([A-Z ]{3,30})",
            r"\b(INDIAN)\b",
        ],
        upper,
    )
    if "INDIAN" in nationality or nationality == "IND":
        nationality = "INDIAN"

    place = _find(
        [
            r"PLACE OF BIRTH[:\s\-]*([A-Z ,\-]{2,40})",
            r"\bPOB[:\s\-]*([A-Z ,\-]{2,40})",
        ],
        upper,
    )

    name = _find(
        [
            r"SURNAME[:\s\-]*([A-Z ]+)\n.*?GIVEN NAME[:\s\-]*([A-Z ]+)",
        ],
        upper,
        flags=re.I | re.S,
    )
    if name:
        # group1 surname group2 given — reconstruct from last match manually
        m = re.search(
            r"SURNAME[:\s\-]*([A-Z ]+).*?GIVEN NAME[S]?[:\s\-]*([A-Z ]+)",
            upper,
            re.S,
        )
        if m:
            name = f"{m.group(2).title().strip()} {m.group(1).title().strip()}".strip()
    if not name:
        m = re.search(r"NAME[:\s\-]*([A-Z ]{3,40})", upper)
        if m:
            name = m.group(1).title().strip()

    # MRZ fallback (last lines often P<IND...)
    mrz_lines = [ln.strip() for ln in upper.splitlines() if re.search(r"[A-Z0-9<]{20,}", ln)]
    if mrz_lines:
        joined = "".join(mrz_lines[-2:])
        m = re.search(r"([A-Z][0-9]{7})", joined)
        if m and not passport_number:
            passport_number = m.group(1)
        m = re.search(r"([0-9]{6})[MF]", joined)
        # YYMMDD DOB in MRZ — leave raw if present
        if m and not dob:
            yymmdd = m.group(1)
            dob = f"YYMMDD:{yymmdd}"

    missing = [
        k
        for k, v in asdict(
            PassportRecord(
                name=name,
                date_of_birth=dob,
                passport_number=passport_number,
                issue_date=issue,
                expiry_date=expiry,
                place_of_birth=place,
                nationality=nationality,
                gender=gender,
            )
        ).items()
        if k not in {"source_file", "confidence_note"} and not v
    ]
    note = "OK" if not missing else "Missing: " + ", ".join(missing)

    return PassportRecord(
        name=name,
        date_of_birth=dob,
        passport_number=passport_number,
        issue_date=issue,
        expiry_date=expiry,
        place_of_birth=place.title() if place else "",
        nationality=nationality,
        gender=gender,
        source_file=source,
        confidence_note=note,
    )


def write_excel(records: list[PassportRecord], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Passports"
    headers = FIELD_HEADERS + ["Source File", "Accuracy Log"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")

    for rec in records:
        ws.append(
            [
                rec.name,
                rec.date_of_birth,
                rec.passport_number,
                rec.issue_date,
                rec.expiry_date,
                rec.place_of_birth,
                rec.nationality,
                rec.gender,
                rec.source_file,
                rec.confidence_note,
            ]
        )
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    wb.save(out_path)


def process_path(input_path: Path, out_path: Path, preview: bool = False) -> list[PassportRecord]:
    images = load_images(input_path)
    if not images:
        raise SystemExit(f"No readable passport files in: {input_path}")

    records: list[PassportRecord] = []
    for source, img in images:
        text = ocr_text(img)
        if preview:
            print(f"\n===== OCR preview: {source} =====\n{text[:1200]}\n")
        rec = parse_indian_passport(text, source)
        records.append(rec)
        print(f"[ok] {Path(source).name}: {rec.passport_number or 'NO_PASSPORT_NO'} | {rec.confidence_note}")

    write_excel(records, out_path)
    print(f"\nExcel written: {out_path.resolve()} ({len(records)} rows)")
    return records


def run_gui() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.title("Passport OCR -> Excel")
    root.geometry("520x220")

    selected = {"path": None}

    info = tk.Label(
        root,
        text="Select PDF/image files or a folder of Indian passport copies.\nOutput Excel columns match the project order.",
        justify="left",
    )
    info.pack(padx=16, pady=12, anchor="w")

    path_var = tk.StringVar(value="(no file selected)")
    tk.Label(root, textvariable=path_var, wraplength=480, justify="left").pack(
        padx=16, anchor="w"
    )

    def choose_file() -> None:
        p = filedialog.askopenfilename(
            title="Select passport file",
            filetypes=[
                ("Passport files", "*.pdf;*.jpg;*.jpeg;*.png;*.tif;*.tiff;*.bmp"),
                ("All files", "*.*"),
            ],
        )
        if p:
            selected["path"] = Path(p)
            path_var.set(p)

    def choose_folder() -> None:
        p = filedialog.askdirectory(title="Select folder")
        if p:
            selected["path"] = Path(p)
            path_var.set(p)

    def run() -> None:
        if not selected["path"]:
            messagebox.showwarning("Missing input", "Please select a file or folder.")
            return
        out = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="passport_results.xlsx",
        )
        if not out:
            return
        try:
            process_path(selected["path"], Path(out), preview=False)
            messagebox.showinfo("Done", f"Saved:\n{out}")
        except Exception as exc:
            messagebox.showerror("Error", str(exc))

    btns = tk.Frame(root)
    btns.pack(pady=16)
    tk.Button(btns, text="Select File", command=choose_file, width=14).grid(row=0, column=0, padx=6)
    tk.Button(btns, text="Select Folder", command=choose_folder, width=14).grid(row=0, column=1, padx=6)
    tk.Button(btns, text="Run OCR", command=run, width=14).grid(row=0, column=2, padx=6)
    root.mainloop()


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Indian passport OCR to Excel")
    parser.add_argument("input", nargs="?", help="Image/PDF file or folder")
    parser.add_argument("--out", default="passport_results.xlsx", help="Output xlsx path")
    parser.add_argument("--preview", action="store_true", help="Print OCR text preview")
    parser.add_argument("--gui", action="store_true", help="Open simple GUI")
    args = parser.parse_args(argv)

    if args.gui or not args.input:
        run_gui()
        return

    process_path(Path(args.input), Path(args.out), preview=args.preview)


if __name__ == "__main__":
    main()
